# Что реализовано:
# Абстрактный класс FileHandler с методами add_items, get_items, delete_items.
# Наследники для JSON, CSV, XLSX, TXT.
# Не перезаписываются данные, добавляются новые вакансии.
# Приватный атрибут файла с именем, есть значение по умолчанию.
# Везде используется remove_duplicates(current, items, key="url").
# Файлы создаются при необходимости (_ensure_file или проверка os.path.exists).
# Данные корректно сохраняются и читаются для всех форматов: JSON, CSV, XLSX, TXT.
# Методы delete_items удаляют элементы по критериям и перезаписывают файл.


import csv
import json
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import DATA_FOLDER
from src.services import remove_duplicates


# ------------------ Абстрактный класс ------------------
class FileHandler(ABC):
    """Абстрактный класс для работы с файлами вакансий."""

    def __init__(self, filename: Optional[str] = None) -> None:
        """:param filename: Имя файла"""
        if filename is None:
            filename = "data/vacancies_data"

        self.__filename: Path = DATA_FOLDER / filename
        Path(self.__filename).parent.mkdir(exist_ok=True, parents=True)

    @abstractmethod
    def add_items(self, items: List[Dict[str, Any]]) -> None: ...

    """Добавляет вакансии в файл."""

    @abstractmethod
    def get_items(self, criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]: ...

    """Возвращает вакансии из файла с возможной фильтрацией."""

    @abstractmethod
    def delete_items(self, criteria: Optional[Dict[str, Any]] = None) -> None: ...

    """Удаляет вакансии из файла по критериям."""

    @property
    def filename(self) -> Path:
        return self.__filename


# ------------------ JSON ------------------
class JSONHandler(FileHandler):
    """работа с JSON-файлом вакансий."""

    def _ensure_file(self) -> None:
        """создаёт файл с заголовком, если его нет."""
        if not Path(self.filename).exists():
            with open(self.filename, "w", encoding="utf-8") as f:
                json.dump([], f, ensure_ascii=False, indent=4)

    def add_items(self, items: List[Dict[str, Any]]) -> None:
        """добавления новых вакансий в файл."""
        self._ensure_file()
        current = self.get_items()
        combined = remove_duplicates(current, items, key="url")
        with open(self.filename, "w", encoding="utf-8") as f:
            json.dump(combined, f, ensure_ascii=False, indent=4)

    def get_items(self, criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """читает вакансии из XLSX и фильтрует их."""
        self._ensure_file()
        with open(self.filename, "r", encoding="utf-8") as f:
            items = json.load(f)
        return _filter_items(items, criteria)

    def delete_items(self, criteria: Optional[Dict[str, Any]] = None) -> None:
        """удаление вакансий из файла"""
        items = self.get_items()
        remaining = _remove_items(items, criteria)
        with open(self.filename, "w", encoding="utf-8") as f:
            json.dump(remaining, f, ensure_ascii=False, indent=4)


# ------------------ CSV ------------------
class CSVHandler(FileHandler):
    def _ensure_file(self) -> None:
        """создаёт файл с заголовком, если его нет."""
        if not Path(self.filename).exists():
            with open(self.filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f, fieldnames=["title", "location", "published_at", "url", "salary", "description"]
                )
                writer.writeheader()

    def add_items(self, items: List[Dict[str, Any]]) -> None:
        """добавления новых вакансий в файл."""
        self._ensure_file()
        current = self.get_items()
        combined = remove_duplicates(current, items, key="url")
        with open(self.filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["title", "location", "published_at", "url", "salary", "description"]
            )
            writer.writeheader()
            writer.writerows(combined)

    def get_items(self, criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """читает вакансии из XLSX и фильтрует их."""
        self._ensure_file()
        with open(self.filename, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            items = list(reader)
        return _filter_items(items, criteria)

    def delete_items(self, criteria: Optional[Dict[str, Any]] = None) -> None:
        """удаление вакансий из файла"""
        items = self.get_items()
        remaining = _remove_items(items, criteria)
        with open(self.filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["title", "location", "published_at", "url", "salary", "description"]
            )
            writer.writeheader()
            writer.writerows(remaining)


class XLSXHandler(FileHandler):
    HEADERS = ["title", "location", "published_at", "url", "salary", "description"]

    def _get_worksheet(self) -> Tuple[Workbook, Worksheet]:
        """возвращает активный лист; гарантирует, что он не None."""
        wb = load_workbook(self.filename)
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Vacancies")
        return wb, ws

    def _ensure_file(self) -> None:
        """создаёт файл с заголовком, если его нет."""
        if not Path(self.filename).exists():
            wb = Workbook()
            ws = wb.active
            if ws is None:  # страховка для mypy
                ws = wb.create_sheet("Vacancies")
            ws.title = "Vacancies"
            ws.append(self.HEADERS)
            wb.save(self.filename)

    def add_items(self, items: List[Dict[str, Any]]) -> None:
        """добавления новых вакансий в файл."""
        self._ensure_file()
        current = self.get_items()
        combined = remove_duplicates(current, items, key="url")
        wb, ws = self._get_worksheet()

        # очищаем все кроме заголовка
        ws.delete_rows(2, ws.max_row)
        for item in combined:
            ws.append([item.get(f, "") for f in self.HEADERS])

        wb.save(self.filename)

    def get_items(self, criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """читает вакансии из XLSX и фильтрует их."""
        self._ensure_file()
        wb, ws = self._get_worksheet()
        rows = list(ws.values)
        if not rows:
            return []
        keys = [str(k) for k in rows[0]]  # гарантируем, что ключи строковые
        items: List[Dict[str, Any]] = [dict(zip(keys, row)) for row in rows[1:]]
        return _filter_items(items, criteria)

    def delete_items(self, criteria: Optional[Dict[str, Any]] = None) -> None:
        """удаление вакансий из файла"""
        items = self.get_items()
        remaining = _remove_items(items, criteria)
        wb, ws = self._get_worksheet()

        # очищаем все кроме заголовка
        ws.delete_rows(2, ws.max_row)
        for item in remaining:
            ws.append([item.get(f, "") for f in self.HEADERS])

        wb.save(self.filename)


# ------------------ TXT ------------------
class TXTHandler(FileHandler):
    def _ensure_file(self) -> None:
        """создаёт файл с заголовком, если его нет."""
        Path(self.filename).touch(exist_ok=True)

    def add_items(self, items: List[Dict[str, Any]]) -> None:
        """добавления новых вакансий в файл."""
        self._ensure_file()
        current = self.get_items()
        combined = remove_duplicates(current, items, key="url")
        with open(self.filename, "w", encoding="utf-8") as f:
            for item in combined:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")

    def get_items(self, criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """читает вакансии из XLSX и фильтрует их."""
        self._ensure_file()
        items = []
        with open(self.filename, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    items.append(json.loads(line.strip()))
        return _filter_items(items, criteria)

    def delete_items(self, criteria: Optional[Dict[str, Any]] = None) -> None:
        """удаление вакансий из файла"""
        items = self.get_items()
        remaining = _remove_items(items, criteria)
        with open(self.filename, "w", encoding="utf-8") as f:
            for item in remaining:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")


# ------------------ Вспомогательные функции ------------------
def _filter_items(items: List[Dict[str, Any]], criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """фильтрации списка вакансий"""
    if not criteria:
        return items
    result = []
    for item in items:
        match = True
        for k, v in criteria.items():
            val = item.get(k)
            if isinstance(v, (list, tuple, set)):
                if str(val) not in map(str, v):
                    match = False
                    break
            else:
                if str(val) != str(v):
                    match = False
                    break
        if match:
            result.append(item)
    return result


def _remove_items(items: List[Dict[str, Any]], criteria: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """удалениt элементов из списка словарей по заданным условиям."""
    if not criteria:
        return []
    remaining = []
    for item in items:
        remove = True
        for k, v in criteria.items():
            val = item.get(k)
            if isinstance(v, (list, tuple, set)):
                if str(val) in map(str, v):
                    continue
                else:
                    remove = False
                    break
            else:
                if str(val) == str(v):
                    continue
                else:
                    remove = False
                    break
        if not remove:
            remaining.append(item)
    return remaining
